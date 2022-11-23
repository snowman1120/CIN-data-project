import re
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def post_processing(Boxes, Texts, save_path, files):
    '''
    This function makes xlsx files from Boxes and Texts.
    1. Loop Boxes and Texts according to document.
    2. In a document, loop boxes and texts according to page.
    3. SR number modification
    4. Consider cell swrap, thin
    5. Save
    '''
    col_cnt = 4
    if len(Texts) > 0:
        col_title = ['SL NO', 'CIN', 'CompanyName', 'Company PAN']
        # col_title = ['SR No','CP No','CA/IA No','Case Purpose','Section','Name of Parties','Remarks','Date of cause list','BENCH','COURT','Applicant Name','Respondent Name']
        pre_rows = 0 # considering multi tables in a page.
        pre_doc_rows = 0
        wb = Workbook()
        ws = wb.active
        ws.title = "new table"
        
        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')) 
        ### loopping every document in multi document ###
        for kk in range(len(Boxes)):
            boxes, index, texts, filename = Boxes[kk][0], np.array(Boxes[kk][1], 'int32'), Texts[kk], files[kk]
            index[0] = 10
            ## Entering file name  ##
            ws.merge_cells(start_row=pre_rows+1, start_column=1, end_row=pre_rows+1, end_column=col_cnt)
            ws.cell(row=pre_rows+1, column=1).value = filename
            ws.cell(row=pre_rows+1, column=1).font = Font(bold=True, size="20")
            ## ------------------  ##
            pre_rows = pre_rows + 2 ## considering sub-title rows. 
            ## thin-border in heading of excel ##
            for j in range(col_cnt):
                ws.cell(row=pre_rows, column=1+j).border = thin_border
                                            
            ### loopping every page in one document ###
            '''
            1. Consider column location change and empty column
            2. Get sub title row number
            3. Set thin border
            4. Insert column title
            '''
            for k in range(len(boxes)):
                box, text = boxes[k], texts[k]
                # if k == 0:
                #     box, text = box[1:], text[1:]
                boxx, text = np.array(box), np.array(text, dtype=str)
                rows = np.unique(boxx[:,0]) # values of y all box((y1,x1,h1,w1), (y2,x2,h2,w2),...)
                cols = np.unique(boxx[:,1]) # values of x all box((y1,x1,h1,w1), (y2,x2,h2,w2),...) 
                ## considering column location change and empty column ##
                inds = [i for i, v in enumerate(index) if v == -1]
                inds.sort(reverse=True)
                for ind in inds:
                    index = np.delete(index, ind)
                cols = cols[index.argsort()]
                inds.sort()
                for ind in inds:
                    cols = np.insert(cols, ind, -1)
                    index = np.insert(index, ind, -1)
                    index.astype(np.uint8)
                
                ## getting sub title row number ##
                sub_tit_row = [] ## this will be include y coor of sub title row 
                for ro in rows:
                    tex = text[np.where(boxx[:,0]==ro)[0]]
                    cnt = 0
                    for te in tex:
                        te = te.lower()
                        if 'cin' in te or 'srn' in te or 'company' in te or 'name' in te:
                            cnt += 1
                        if cnt > 1: 
                            sub_tit_row.append(ro)
                            break
                                
                for i in range(len(rows)-len(sub_tit_row)):
                    # ws = heading_insert(ws, pre_rows+i, heading)     
                    for j in range(col_cnt):
                        ws.cell(row=pre_rows+1+i, column=1+j).border = thin_border
                                    
                ### loopping every element in one one page ###
                '''
                1. Get number of row and column of element.
                2. Consider merged cell.
                3. Modify texts
                '''
                subtit_check = False
                for i, bo in enumerate(boxx):
                    y,x,h,w = bo 
                    text_val = text[i]
                    
                    if y in sub_tit_row:
                        subtit_check = True
                        continue
                    if subtit_check: 
                        pre_rows += -1
                        subtit_check = False
                    row_num = np.where(rows==y)[0][0] # number of row in table
                    try:
                        col_num = np.where(cols==x)[0][0] # number of column in table
                    except:
                        pass

                    # Consider merged cell.
                    r_cnt = 0
                    for j in range(row_num+1, len(rows)):
                        if rows[j] < y + h*0.98: 
                            r_cnt = r_cnt + 1
                    c_cnt = 0
                    try:
                        text_val = text_val.upper()
                        if len(text_val) > 21:
                            text_val = re.sub('(JJH|J/H)', 'JH', text_val)                        
                        # Text modification
                        if r_cnt == 0 and c_cnt == 0: # if merged cell is not
                            if col_num == 1: ## CIN column
                                strp_chars = "^#;$`'-_=*,‘:¢ \n"
                                text_val = text_val.strip(strp_chars)
                                text1, text2, text3, text4, text5, text6 = text_val[0], \
                                        text_val[1:6], text_val[6:8], text_val[8:12], text_val[12:15], text_val[15:21]
                                # 1:alpha, 2:numeric, 3:alpha, 4:numeric, 5:alpha, 6:numeric
                                text_val = CIN_modified_text(text1, text2, text3, text4, text5, text6)
                            elif col_num == 3: ## Pan Num column
                                text1, text2, text3 = text_val[0:5], text_val[5:9], text_val[9]
                                strp_chars = "^#;$`'-_=*,‘:¢ \n"
                                text_val = text_val.strip(strp_chars)
                                text_val = PAN_modified_text(text1, text2, text3)
                            #     if len(text_val) == 1:
                            #         text_val = ""
                            #     else:
                            #         text_val = re.sub('(Appeali|Appeall)', 'Appeal/', text_val)
                            #         text_val = re.sub('(/88/|/B8/|/8B/)', '/BB/', text_val)
                            #         text_val = re.sub('(CP(18)|CP(1B))', 'CP(IB)', text_val)          
                            #         text_val = re.sub('(201B|20lB|20IB)', '2018', text_val)
                            #         text_val = text_val.replace('cP', 'CP')
                                
                            ws.cell(row=row_num+pre_rows+1, column=col_num+1).value = text_val
                        else:  
                            ws.merge_cells(start_row=row_num+pre_rows+1, start_column=col_num+1, end_row=row_num+pre_rows+r_cnt+1, end_column=col_num+c_cnt+1)
                            ws.cell(row=row_num+pre_rows+1, column=col_num+1).value = text_val
                    except:
                        pass
                pre_rows = pre_rows + len(rows) - len(sub_tit_row)
                if k == 0: pre_rows += 1
            pre_rows = pre_rows + 1
            # column title insert
            for i in range(len(col_title)):
                ws.cell(pre_doc_rows+2,i+1).value = col_title[i]
                ws.cell(pre_doc_rows+2,i+1).font = Font(bold=True)

            pre_doc_rows = pre_rows
        
        # SR No modification
        sr_no, new_sr = [], []
        for i in range(2, ws.max_row+1):
            val = ws.cell(i, 1).value
            
            if val is not None:
                strp_chars = "|^#;$`'-_=\/*‘., \n"
                val = val.strip(strp_chars)            
            try:
                sr_no.append(int(val))
            except:
                sr_no.append(val)

        for i in range(len(sr_no)):
            val = get_nei(i, sr_no)
            sr_no[i] = val
            new_sr.append(val)
        ## put new sr no 
        for i in range(2, ws.max_row+1):
            val = new_sr[i-2]
            if val is not None:
                ws.cell(i, 1).value = val

        # cell swrap, thin
        row_no = 1
        for i in ws.rows:
            for j in range(len(i)):
                ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            row_no = row_no + 1  

        # column width
        column_width = [6, 30, 50, 25]
        for i in range(col_cnt):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
        ws.sheet_view.zoomScale = 75
        # save
        wb.save(save_path)
    else:
        print("=== Table of this pdf is not detected ===")

    return None

def get_nei(i, sr_no):
    val_i = sr_no[i]
    if val_i != 'SL NO' and val_i !='' and val_i is not None:
        # get pre_num and next_num #
        pre_num, next_num = None, None
        for j in range(1, len(sr_no)):
            if i-j >= 0:
                if sr_no[i-j] == 'SL NO':
                    break
                elif isinstance(sr_no[i-j], int):
                    pre_num = sr_no[i-j]
                    break
        try:
            if sr_no[i+1] == 'SL NO':
                return val_i
        except:
            pass

        for j in range(1, len(sr_no)):
            if i+j < len(sr_no):
                if isinstance(sr_no[i+j], str):
                    break
                elif isinstance(sr_no[i+j], int):
                    next_num = sr_no[i+j]
                    break
        ## #######################################
        
        if next_num is None and pre_num is None:
            val = val_i
        elif next_num is None:
            # if isinstance(val_i, int) and (int(str(val_i)[-1]) - int(str(pre_num)[-1]) == 1):
            #     val = pre_num + 1
            # else: val = val_i
            val = pre_num + 1
        elif pre_num is None:
            if next_num == 2:
                val = 1
            else:
                val = val_i
        else:
            try:
                if val_i - pre_num == 1 or next_num - val_i == 1:
                    val = val_i
                elif next_num - pre_num ==2:
                    val = pre_num + 1
                else: val = next_num - 1
            except:
                if next_num - pre_num ==2:
                    val = pre_num + 1
                else: val = next_num - 1                
    else:
        val = val_i
    return val
def CIN_modified_text(text1, text2, text3, text4, text5, text6):
    ## 0 and O ##
    ## 5 and S ##
    ## J and ) ##
    ## 7 and T ##
    ### Numeric -> Alphabet ###
    if '0' in text1: text1 = text1.replace('0', 'O')
    if '0' in text3: text3 = text3.replace('0', 'O')
    if '0' in text5: text5 = text5.replace('0', 'O')
    if '5' in text1: text1 = text1.replace('5', 'S')    
    if '5' in text3: text3 = text3.replace('5', 'S')
    if '5' in text5: text5 = text5.replace('5', 'S')  
    if ')' in text1: text1 = text1.replace(')', 'J')    
    if ')' in text3: text3 = text3.replace(')', 'J')
    if ')' in text5: text5 = text5.replace(')', 'J')  
    ### Alphabet -> Numeric ###
    if 'O' in text2: text2 = text2.replace('O', '0')
    if 'O' in text4: text4 = text4.replace('O', '0')
    if 'O' in text6: text6 = text6.replace('O', '0')   
    if 'S' in text2: text2 = text2.replace('S', '5')
    if 'S' in text4: text4 = text4.replace('S', '5')
    if 'S' in text6: text6 = text6.replace('S', '5')
    if 'T' in text2: text2 = text2.replace('T', '7')
    if 'T' in text4: text4 = text4.replace('T', '7')
    if 'T' in text6: text6 = text6.replace('T', '7')

    return ''.join([text1, text2, text3, text4, text5, text6])

def PAN_modified_text(text1, text2, text3):

    if '0' in text1: text1 = text1.replace('0', 'O')
    if '0' in text3: text3 = text3.replace('0', 'O')
    if '5' in text1: text1 = text1.replace('5', 'S')    
    if '5' in text3: text3 = text3.replace('5', 'S')
    if ')' in text1: text1 = text1.replace(')', 'J')    
    if ')' in text3: text3 = text3.replace(')', 'J')

    if 'O' in text2: text2 = text2.replace('O', '0')
    if 'S' in text2: text2 = text2.replace('S', '5')
    if 'T' in text2: text2 = text2.replace('T', '7') 

    return ''.join([text1, text2, text3])